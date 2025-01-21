import streamlit as st
import pandas as pd
import base64
import matplotlib.pyplot as plt
import sqlite3
import hashlib
from PIL import Image
import pickle
from sklearn.calibration import calibration_curve
from sklearn.metrics import brier_score_loss, accuracy_score, recall_score, precision_score, f1_score, roc_auc_score, roc_curve
from sklearn.metrics import roc_curve, auc, accuracy_score, recall_score, precision_score, f1_score 
from pycaret import classification
from pycaret.classification import setup, get_logs,create_model, predict_model, tune_model, save_model,load_model
import shap
shap.initjs()
import joblib
from sklearn.preprocessing import OneHotEncoder
from sklearn.tree import DecisionTreeClassifier
import numpy as np
import streamlit.components.v1 as components
import openpyxl
import os
from sklearn.ensemble import RandomForestClassifier

class DynamicWeightedForest:
    def __call__(self, X):
        return self.predict_proba(X)
                
    def __init__(self, base_trees):
        self.trees = base_trees
        self.tree_weights = np.ones(len(self.trees)) / len(self.trees)

    def predict_proba(self, X):
        weighted_votes = np.zeros((X.shape[0], 2))  # Ensure the output is always (n_samples, 2)
        
        for i, tree in enumerate(self.trees):
            proba = tree.predict_proba(X)
            
            print(f"Tree {i} proba shape: {proba.shape}")
            print(f"Tree {i} proba: {proba}") 
        
            if proba.shape[1] == 1:
                proba = np.hstack([1 - proba, proba])  # Add missing class probability
            
            weighted_votes += self.tree_weights[i] * proba

        return weighted_votes / np.sum(self.tree_weights)


    def get_weighted_shap_values(self, X):
        shap_values_sum = np.zeros((X.shape[0], X.shape[1]))
        expected_value_sum = 0
        for tree, weight in zip(self.trees, self.tree_weights):
            explainer = shap.TreeExplainer(tree)
            shap_values_tree = explainer.shap_values(X)[1]  # Get the SHAP values for the positive class
            print(f"Tree {i} SHAP values: {shap_values_tree}")
            expected_value_tree = explainer.expected_value[1]
            shap_values_sum += weight * shap_values_tree
            expected_value_sum += weight * expected_value_tree
        print(f"Final SHAP values sum: {shap_values_sum}")
        print(f"Final expected value sum: {expected_value_sum}")
        return shap_values_sum, expected_value_sum


    def update_weights(self, X, y):
        for i, tree in enumerate(self.trees):
            predictions = tree.predict(X)
            accuracy = np.mean(predictions == y)
            self.tree_weights[i] = accuracy
        self.tree_weights /= np.sum(self.tree_weights)
        print(f"Updated tree weights: {self.tree_weights}") 

    def add_tree(self, new_tree):
        self.trees.append(new_tree)
        self.tree_weights = np.append(self.tree_weights, [1.0])
        self.tree_weights /= np.sum(self.tree_weights)
    
        # Debugging: Check the shape after adding a new tree
        print(f"After adding new tree, number of trees: {len(self.trees)}")

    def save_model(self, model_name):
        st.write(f"Saving model to {model_name}")  
        joblib.dump(self, model_name)
            
    @staticmethod
    def load_model(model_name):
        st.write(f"Loading model from {model_name}")  
        if os.path.exists(model_name):
            return joblib.load(model_name)
        else:
            st.error(f"Model file {model_name} not found.")
            return None

def load_global_model():
    model_file = 'global_weighted_forest.pkl'  # 术前模型
    st.write(f"Attempting to load global model: {model_file}")
    try:
        if os.path.exists(model_file):
            try:
                model = DynamicWeightedForest.load_model(model_file)
                st.write(f"Model loaded successfully: {model_file}")
                return model
            except EOFError:
                st.error(f"Model file is corrupted: {model_file}. Deleting and regenerating...")
                os.remove(model_file)
        else:
            st.warning(f"Model file not found: {model_file}. Creating a new model.")
        
        # 加载初始模型
        initial_model = joblib.load('tuned_rf_pre_BUN.pkl')
        st.write("Initialized a new model from base trees.")
        return DynamicWeightedForest(initial_model.estimators_)
    except Exception as e:
        st.error(f"Failed to load or create global model: {e}")
        return None

def load_global_model2():
    model_file = 'global_weighted_forest2.pkl'  # 术中模型
    st.write(f"Attempting to load global model: {model_file}")
    try:
        if os.path.exists(model_file):
            try:
                model = DynamicWeightedForest.load_model(model_file)
                st.write(f"Model loaded successfully: {model_file}")
                return model
            except EOFError:
                st.error(f"Model file is corrupted: {model_file}. Deleting and regenerating...")
                os.remove(model_file)
        else:
            st.warning(f"Model file not found: {model_file}. Creating a new model.")
        
        # 加载术中初始模型
        initial_model = joblib.load('tuned_rf_intra_BUN.pkl')
        st.write("Initialized a new model from base trees.")
        return DynamicWeightedForest(initial_model.estimators_)
    except Exception as e:
        st.error(f"Failed to load or create global model: {e}")
        return None

def get_db_connection():
    conn = sqlite3.connect('users.db')
    conn.row_factory = sqlite3.Row
    return conn

def initialize_database():
    conn = get_db_connection()
    conn.execute('''
        CREATE TABLE IF NOT EXISTS users (
            username TEXT NOT NULL UNIQUE,
            password TEXT NOT NULL
        )
    ''')
    conn.commit()
    conn.close()

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def register_user(username, password):
    conn = get_db_connection()
    try:
        conn.execute("INSERT INTO users (username, password) VALUES (?, ?)",
                     (username, hash_password(password)))
        conn.commit()
        return True
    except sqlite3.IntegrityError:
        return False
    finally:
        conn.close()

def validate_user(username, password):
    conn = get_db_connection()
    user = conn.execute("SELECT * FROM users WHERE username = ? AND password = ?",
                        (username, hash_password(password))).fetchone()
    conn.close()
    return user is not None

def login_page():
    st.title("User Login")
    
    with st.form("login_form"):
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")
        submitted = st.form_submit_button("Login")

        if submitted:
            if validate_user(username, password):
                st.session_state['is_logged_in'] = True
                st.success("Login successful!")
                st.experimental_rerun()
            else:
                st.error("Invalid username or password.")

def register_page():
    st.title("User Registration")

    with st.form("register_form"):
        new_username = st.text_input("New Username")
        new_password = st.text_input("New Password", type="password")
        register_button = st.form_submit_button("Register")

        if register_button:
            if register_user(new_username, new_password):
                st.success("Registration successful! Please log in.")
            else:
                st.error("That username is already taken. Please try another username.")
                
def prediction_page():
    import streamlit as st

    # Set the page title
    st.set_page_config(page_title="My Application", layout="centered")


    page = st.sidebar.selectbox("Select a Page", ["Home Page", "Prediction"])

    if page == "Home Page":
        st.title("Welcome to Functional outcome prediction App for patients with posterior circulation large vessel occlusion after mechanical thrombectomy")
    
        st.header("Summary")
        st.write("""
            This application aims to predict functional outcome in  patients with posterior circulation large vessel occlusion following mechanical thrombectomy，thus facilitates informed clinical judgment, supports personalized treatment and follow-up plans, and establishes realistic treatment expectations.
        """)

        st.header("Main Features")
        features = [
            "✔️ Implementation of preoperative, intraoperative, and postoperative prediction models to dynamically update predictions of functional outcomes.",
            "✔️ Support for batch predictions of functional outcomes for multiple patients.",
            "✔️ Ability to predict outcomes for patients with missing variable values.",
            "✔️ Facilitation of the interpretation of how the model provides personalized predictions for specific cases.",
            "✔️ Consideration of changing environments with automatic deployment of updated prediction models."
        ]
        for feature in features:
            st.write(feature)

        st.header("How to Use")
        st.markdown("""
            To the left, is a dropdown main menu for navigating to each page in the present App:<br><br>
            
            &bull; **Home Page:** We are here!<br>
            
            &bull; **Prediction:** Overview of the prediction section.<br>
            
            &bull; **Preoperative_number:** Manage preoperative predictions by inputting the necessary data.<br>
            
            &bull; **Preoperative_batch:** Process preoperative batch predictions by uploading a file.<br>
            
            &bull; **Perioperative_number:** Manage perioperative predictions by inputting the necessary data.<br>
            
            &bull; **Perioperative_batch:** Process perioperative batch predictions by uploading a file.<br>
            
            &bull; **Postoperative_number:** Manage postoperative predictions by inputting the necessary data.<br>
            
            &bull; **Postoperative_batch:** Process postoperative batch predictions by uploading a file.<br>
            
            """, unsafe_allow_html=True)

        pdf_file_path = r"123.pdf"
        if os.path.exists(pdf_file_path):
            st.markdown("Click here to download the manual for more detailed usage instructions:")
            with open(pdf_file_path, "rb") as f:
                st.download_button(label="User Manual.pdf",data=f, file_name="User Manual.pdf",mime="application/pdf" )
        else:
            st.error("指定的文件不存在，请检查文件路径。")
        
        st.header("Contact Us")
        st.write("""
            If you have any questions, please contact the support team:
            - Email: 2894683001@qq.com
            
        """)
        
        st.header("Useful Links")
        st.markdown(
        """
        An app designed to predict functional outcomes for patients with anterior circulation large vessel occlusion following mechanical thrombectomy:
         - [Visit our partner site](https://zhelvyao-123-60-anterior.streamlit.app/)
         """)

        st.markdown(
            f"""
            <style>
                .appview-container .main .block-container{{
                    max-width: {1150}px;
                    padding-top: {5}rem;
                    padding-right: {10}rem;
                    padding-left: {10}rem;
                    padding-bottom: {5}rem;
                }}
                .reportview-container .main {{
                    color: white;
                    background-color: black;
                }}
            </style>
            """,
            unsafe_allow_html=True
        )
    
        image = Image.open('it.tif')
        st.image(image, use_column_width=True)


    elif page == "Prediction":
        st.title('Functional outcome prediction App for patients with posterior circulation large vessel occlusion after mechanical thrombectomy')
    
        model = joblib.load('tuned_rf_pre_BUN.pkl')
        model2 = load_model('tuned_rf_pre_BUN_model')
        model3 = joblib.load('tuned_rf_intra_BUN.pkl')
        model4 = load_model('tuned_rf_intra_BUN_model')
        model5 = joblib.load('tuned_rf_post_BUN.pkl')
        model6 = load_model('tuned_rf_post_BUN_model')
        
        if st.button('Reset to Initial Model'):
            # Reset everything and load the initial model directly
            st.session_state['new_data'] = pd.DataFrame()  # 清空增量学习的数据
            current_model = joblib.load('tuned_rf_pre_BUN.pkl')  # 直接加载初始模型
            st.session_state['current_model'] = current_model  # 将初始模型存储到 session state
            st.success("Model has been reset to the initial model!")
        else:
            # 使用之前加载的模型
            if 'current_model' in st.session_state:
                current_model = st.session_state['current_model']
                st.write(f"Using model from session state: {type(current_model)}")  # 打印模型类型，确认是正确的模型
            else:
                current_model = load_global_model()  # 加载初始模型
                st.write("Model loaded as no model was found in session state.")  # 输出模型加载信息

        def st_shap(plot):
            shap_html = f"<head>{shap.getjs()}</head><body>{plot.html()}</body>"
            components.html(shap_html)

        prediction_type = st.sidebar.selectbox(
            "How would you like to predict?",
            ("Preoperative_number", "Preoperative_batch", "Intraoperative_number", "Intraoperative_batch", 
             "Postoperative_number", "Postoperative_batch")
        )
    
                
        if prediction_type == "Preoperative_number":
            st.subheader("Preoperative Number Prediction")
            st.write("Please fill in the blanks with corresponding data.")
            
            NIHSS = st.number_input('NIHSS', min_value = 4, max_value = 38, value = 10) 
            GCS = st.number_input('GCS', min_value = 0, max_value = 15 , value = 10) 
            pre_eGFR = st.number_input('pre_eGFR', min_value = 10.00, max_value = 250.00, value = 111.5)
            pre_glucose = st.number_input('pre_glucose', min_value = 2.50, max_value = 25.00, value = 7.78)
            PC_ASPECTS = st.number_input('PC_ASPECTS', min_value = 0.0, max_value = 10.0, value = 8.0)
            Age = st.number_input('Age', min_value = 0, max_value = 120, value = 60)
            pre_BUN = st.number_input('pre_BUN', min_value = 0.00, max_value = 30.00, value = 10.20)
        
            features = { 
                'NIHSS': NIHSS, 
                'GCS': GCS, 
                'pre_eGFR': pre_eGFR,
                'pre_glucose': pre_glucose, 
                'PC_ASPECTS': PC_ASPECTS, 
                'Age': Age, 
                'pre_BUN': pre_BUN
            }
        
            input_df = pd.DataFrame([features])
        
            if 'new_data' not in st.session_state:
                st.session_state['new_data'] = pd.DataFrame(columns=input_df.columns.tolist() + ['label'])
        
            # Prediction logic
            if st.button('Predict'):
                try:
                    input_array = input_df.values.reshape(1, -1)
            
                    # For RandomForestClassifier
                    if isinstance(current_model, RandomForestClassifier):
                        output = current_model.predict_proba(input_array)
            
                        if output.shape[1] == 1:
                            st.warning("The model seems to predict only one class. Adding probabilities for the missing class.")
                            output = np.hstack([1 - output, output])
            
                        probability = output[:, 1]
            
                        # SHAP for RandomForestClassifier
                        explainer = shap.TreeExplainer(current_model)
                        shap_values = explainer.shap_values(input_array)
                        expected_value = explainer.expected_value[1]
            
                    # For DynamicWeightedForest
                    if isinstance(current_model, DynamicWeightedForest):
                        # Check if there are any trees in the model
                        if len(current_model.trees) == 0:
                            st.warning("No trees found in the DynamicWeightedForest model!")
                        
                        output = current_model.predict_proba(input_array)
                        
                        # Ensure the output has the expected shape and is valid
                        if output.shape[1] == 1:
                            st.warning("The model seems to predict only one class. Adding probabilities for the missing class.")
                            output = np.hstack([1 - output, output])
                    
                        probability = output[:, 1]
                    
                        # SHAP for DynamicWeightedForest
                        shap_values, expected_value = current_model.get_weighted_shap_values(input_array)
                        
                        # Debugging: Check the output of the DynamicWeightedForest model
                        print(f"Incremental learning model output: {output}")
                        print(f"SHAP values: {shap_values}")
                        print(f"Expected value: {expected_value}")
                    
                        # Visualize SHAP values using force plot
                        st_shap(shap.force_plot(expected_value, shap_values, input_array))
                    
                        # Ensure shap_values is 1D for visualization
                        if isinstance(shap_values, list):
                            shap_values = shap_values[1]  # Use the SHAP values for the positive class (index 1)
                        elif isinstance(shap_values, np.ndarray):
                            shap_values = shap_values.flatten()  # Flatten to ensure it's 1D
                    
                        st.write(f"Flattened SHAP values: {shap_values}")
                    
                        shap_values_flat = shap_values.flatten()
                        shap_df = pd.DataFrame({'Feature': input_df.columns, 'SHAP Value': shap_values_flat})
                        st.write("SHAP values for each feature:")
                        st.dataframe(shap_df)

        
                except Exception as e:
                    st.error(f"Error during prediction: {e}")

        
            # Adding data for Incremental Learning
            if st.button('Add Data for Learning'):
                try:
                    label = int(st.selectbox('Outcome for Learning', [0, 1]))  # Ensure this is inside the button's block
        
                    # Add label to the input data
                    new_data = input_df.copy()
                    new_data['label'] = label
                    st.session_state['new_data'] = pd.concat([st.session_state['new_data'], new_data], ignore_index=True)
        
                    accumulated_data = st.session_state['new_data']
                    X = accumulated_data.drop(columns=['label'])
                    y = accumulated_data['label'].astype(int)
        
                    st.write("Accumulated training data preview:")
                    st.dataframe(accumulated_data)
                    st.write(f"Features shape: {X.shape}, Labels shape: {y.shape}")
                    st.write(f"Unique labels in training data: {y.unique()}")
        
                    if isinstance(current_model, RandomForestClassifier):
                        current_model.fit(X, y)
                        joblib.dump(current_model, 'tuned_rf_pre_BUN.pkl')  # Save the updated model
                        st.success("RandomForestClassifier model updated successfully!")
        
                    elif isinstance(current_model, DynamicWeightedForest):
                        new_tree = DecisionTreeClassifier(random_state=42)
                        new_tree.fit(X, y)
                        current_model.add_tree(new_tree)
                        current_model.update_weights(X, y)
                        current_model.save_model('global_weighted_forest.pkl')  # Save the updated DWF model
                        st.success("DynamicWeightedForest model updated successfully!")
        
                except Exception as e:
                    st.error(f"Error during model update: {e}")

            
        elif prediction_type == "Preoperative_batch":
            st.subheader("Preoperative Batch Prediction")
            def plot_roc_curve(y_true, y_scores): 
                fpr, tpr, thresholds = roc_curve(y_true, y_scores) 
                roc_auc = auc(fpr, tpr)

                plt.figure(figsize=(10, 6)) 
                plt.plot(fpr, tpr, color='blue', label='ROC curve (area = %0.2f)' % roc_auc) 
                plt.plot([0, 1], [0, 1], color='red', linestyle='--') 
                plt.xlim([0.0, 1.0]) 
                plt.ylim([0.0, 1.05]) 
                plt.xlabel('False Positive Rate') 
                plt.ylabel('True Positive Rate')
                plt.title('Receiver Operating Characteristic (ROC)') 
                plt.legend(loc='lower right') 
                plt.grid() 
                st.pyplot(plt) 
            st.write("This section will handle preoperative batch predictions.Please click on the link to download the form and fill in the corresponding data. After that, click on the Browse files button to upload file for prediciton, you can see the prediction of the classifier at the bottom. This page supports batch prediction of the outcome of multiple patients at one time, and can predict the outcome of patients with missing values.")

            csv_exporter = openpyxl.Workbook()
            sheet = csv_exporter.active
            sheet.cell(row=1, column=1).value = 'NIHSS'
            sheet.cell(row=1, column=2).value = 'GCS'
            sheet.cell(row=1, column=3).value = 'pre_eGFR'
            sheet.cell(row=1, column=4).value = 'pre_glucose'
            sheet.cell(row=1, column=5).value = 'PC_ASPECTS'
            sheet.cell(row=1, column=6).value = 'Age'
            sheet.cell(row=1, column=7).value = 'pre_BUN'
            csv_file_name = 'for_predictions.csv'
            csv_exporter.save(csv_file_name)
            if os.path.exists(csv_file_name):
                data = open(csv_file_name, 'rb').read()
                b64 = base64.b64encode(data).decode('UTF-8')
                href = f'<a href="data:file/csv;base64,{b64}" download="{csv_file_name}">Download csv file</a>'
                st.markdown(href, unsafe_allow_html=True)
            else:
                 st.error(f"文件 '{csv_file_name}' 找不到，请检查文件生成的过程。")
            csv_exporter.close()

            file_upload = st.file_uploader("Upload CSV file for predictions", type=["csv"])

            if file_upload is not None: 
                try: 
                    data = pd.read_csv(file_upload, sep=',', error_bad_lines=False) 

                    if 'MRSI' in data.columns: 
                        y_true = data['MRSI'].values 
                        predictions = model2.predict_proba(data)[:, 1] 
                        predictions_df = pd.DataFrame(predictions, columns=['Predictions']) 
                        st.write(predictions)
                        result_data = data.copy() 
                        result_data['Predictions'] = predictions_df 
                        result_file_path = 'predictions_with_results.csv' 
                        result_data.to_csv(result_file_path, index=False) 
 
                        with open(result_file_path, 'rb') as f: 
                            output_data = f.read()
                            b64 = base64.b64encode(output_data).decode('UTF-8')
                            download_link = f'<a href="data:file/csv;base64,{b64}" download="predictions_with_results.csv">Download predictions with results</a>'
                            st.markdown(download_link, unsafe_allow_html=True)

                        add_data = st.selectbox('Outcome for Learning', [0, 1])
                        if st.button('Add Data for Learning'): 
                            X = data.drop(columns=['MRSI']) 
                            y = data['MRSI'] 
                            rf_model2 = model2.named_steps['trained_model']
                            pre_weighted_forest2 = DynamicWeightedForest(rf_model2.estimators_)
                            new_tree = DecisionTreeClassifier(random_state=42)
                            new_tree.fit(X, y) 
                            pre_weighted_forest2.add_tree(new_tree)
                            pre_weighted_forest2.update_weights(X, y) 
                            st.success("New tree added and weights updated dynamically!")

                        def plot_combined_graphs(y_true, y_scores):
                            fig, axs = plt.subplots(1, 2, figsize=(14, 6))
                            fpr, tpr, _ = roc_curve(y_true, y_scores)
                            roc_auc = auc(fpr, tpr)
                            axs[0].plot(fpr, tpr, color='blue', label='ROC curve (area = %0.2f)' % roc_auc)
                            axs[0].plot([0, 1], [0, 1], color='red', linestyle='--')
                            axs[0].set_xlim([0.0, 1.0])
                            axs[0].set_ylim([0.0, 1.05])
                            axs[0].set_xlabel('False Positive Rate')
                            axs[0].set_ylabel('True Positive Rate')
                            axs[0].set_title('Receiver Operating Characteristic (ROC)')
                            axs[0].legend(loc='lower right')
                            axs[0].grid()

                            prob_true, prob_pred = calibration_curve(y_true, y_scores, n_bins=10)
                            axs[1].plot(prob_pred, prob_true, marker='o', label='Calibrated Model', color='b')
                            axs[1].plot([0, 1], [0, 1], linestyle='--', label='Perfectly Calibrated', color='r')
                            axs[1].set_title('Brier Score Calibration Plot')
                            axs[1].set_xlabel('Mean Predicted Probability')
                            axs[1].set_ylabel('Fraction of Positives')
                            axs[1].set_xlim([0, 1])
                            axs[1].set_ylim([0, 1])
                            axs[1].legend()
                            axs[1].grid()
                            st.pyplot(fig)

                        if len(data) >= 10:
                            y_pred = (predictions > 0.5).astype(int)
                            accuracy = accuracy_score(y_true, y_pred)
                            recall = recall_score(y_true, y_pred)
                            precision = precision_score(y_true, y_pred)
                            f1 = f1_score(y_true, y_pred)
                            roc_auc = auc(*roc_curve(y_true, predictions)[:2])

                            st.write(f"Accuracy: {accuracy:.2f}")
                            st.write(f"Recall: {recall:.2f}")
                            st.write(f"Precision: {precision:.2f}")
                            st.write(f"F1 Score: {f1:.2f}")
                            st.write(f"AUC: {roc_auc:.2f}")
                            brier_score = brier_score_loss(y_true, predictions)
                            st.write(f"Brier Score: {brier_score:.2f}")

                            plot_combined_graphs(y_true, predictions)

                        else:
                            st.warning("Not enough samples for ROC curve plotting. Please upload at least 10 samples.") 

                            st.write(predictions_df) 
                    else:                      
                        predictions = model2.predict_proba(data)[:,1] 
                        predictions = pd.DataFrame(predictions,columns = ['Predictions'])
                        st.write(predictions)
                        result_data = data.copy() 
                        result_data['Predictions'] = predictions 
                        result_file_path = 'predictions_with_results.csv'
                        result_data.to_csv(result_file_path, index=False)
                        with open(result_file_path, 'rb') as f:
                            data = f.read()
                            b64 = base64.b64encode(data).decode('UTF-8')
                            download_link = f'<a href="data:file/csv;base64,{b64}" download="predictions_with_results.csv">Download predictions with results</a>' 
                            st.markdown(download_link, unsafe_allow_html=True)

                except Exception as e: 
                    st.error(f"Error reading the CSV file: {e}") 
                    
        elif prediction_type == "Intraoperative_number":
            st.subheader("Intraoperative Number Prediction")
            st.write("This section will handle intraoperative number predictions.please fill in the blanks with corresponding data. After that,click on the Predict button at the bottom to see the prediction of the classifier.")

            NIHSS = st.number_input('NIHSS', min_value = 4,max_value = 38,value = 10) 
            GCS= st.number_input('GCS', min_value = 0,max_value = 15 ,value = 10) 
            pre_eGFR = st.number_input('pre_eGFR', min_value = 10.00,max_value = 250.00,value = 111.5)
            PC_ASPECTS = st.number_input('PC_ASPECTS', min_value = 0.0,max_value = 10.0,value = 8.0)
            Age = st.number_input('Age', min_value = 0,max_value = 120,value = 60)
            pre_BUN = st.number_input('pre_BUN', min_value = 0.20,max_value = 30.00,value = 3.20)
            procedural_time = st.number_input('procedural time', min_value=0.00, max_value=350.00, value=60.00)          
            output = ""

            features = { 
                'NIHSS': NIHSS, 
                'GCS': GCS, 
                'pre_eGFR': pre_eGFR, 
                'PC_ASPECTS': PC_ASPECTS,
                'Age': Age,
                'pre_BUN': pre_BUN,
                'procedural time': procedural_time                  
               
                 }

            print(features) 

            input_df = pd.DataFrame([features]) 
            print(input_df) 
            if 'new_data' not in st.session_state:
                st.session_state['new_data'] = pd.DataFrame(columns=input_df.columns.tolist() + ['label'])

            if st.button('Predict'):
                try:
                    output = current_model.predict_proba(input_df)  # 术前模型的预测

                    if output.shape[1] == 1:  # 只有一个类的概率
                        st.warning("The model seems to predict only one class. Adding probabilities for the missing class.")
                        output = np.hstack([1 - output, output])  # 补充缺失的类别

                    probability = output[:, 1]
                    explainer = shap.Explainer(current_model)
         
                    st_shap(shap.force_plot(expected_value, shap_values[0], input_df))

                    shap_df = pd.DataFrame({
                        'Feature': input_df.columns,
                        'SHAP Value': shap_values[0]
                    })
                    st.write("SHAP values for each feature:")
                    st.dataframe(shap_df)

                except Exception as e:
                    st.error(f"Error during prediction: {e}")
        
            label = int(st.selectbox('Outcome for Learning', [0, 1]))
        
            if st.button('Add Data for Learning'):
                try:
                    new_data = input_df.copy()
                    new_data['label'] = label
                    st.session_state['new_data'] = pd.concat([st.session_state['new_data'], new_data], ignore_index=True)
        
                    accumulated_data = st.session_state['new_data']
                    X = accumulated_data.drop(columns=['label'])
                    y = accumulated_data['label'].astype(int)  

                    st.write("Accumulated training data preview:")
                    st.dataframe(accumulated_data)
                    st.write(f"Features shape: {X.shape}, Labels shape: {y.shape}")
                    st.write(f"Unique labels in training data: {y.unique()}")
        
                    new_tree = DecisionTreeClassifier(random_state=42)
                    new_tree.fit(X, y)
        
                    current_model2.add_tree(new_tree)
                    current_model2.update_weights(X, y)
                    current_model2.save_model('global_weighted_forest2.pkl')
        
                    st.success("New tree added and weights updated dynamically with accumulated data!")
        
                except Exception as e:
                    st.error(f"Error during model update: {e}")

        elif prediction_type == "Intraoperative_batch":
            st.subheader("Intraoperative Batch Prediction")
            st.write("This section will handle intraoperative batch predictions.Please click on the link to download the form and fill in the corresponding data. After that, click on the Browse files button to upload file for prediciton, you can see the prediction of the classifier at the bottom. This page supports batch prediction of the outcome of multiple patients at one time, and can predict the outcome of patients with missing values.")
            def plot_roc_curve(y_true, y_scores): 
                fpr, tpr, thresholds = roc_curve(y_true, y_scores) 
                roc_auc = auc(fpr, tpr)

                plt.figure(figsize=(10, 6)) 
                plt.plot(fpr, tpr, color='blue', label='ROC curve (area = %0.2f)' % roc_auc) 
                plt.plot([0, 1], [0, 1], color='red', linestyle='--') 
                plt.xlim([0.0, 1.0]) 
                plt.ylim([0.0, 1.05]) 
                plt.xlabel('False Positive Rate') 
                plt.ylabel('True Positive Rate')
                plt.title('Receiver Operating Characteristic (ROC)') 
                plt.legend(loc='lower right') 
                plt.grid() 
                st.pyplot(plt) 

            csv_exporter = openpyxl.Workbook()
            sheet = csv_exporter.active
            sheet.cell(row=1,column=1).value='NIHSS'
            sheet.cell(row=1,column=2).value='GCS'
            sheet.cell(row=1,column=3).value='pre_eGFR'
            sheet.cell(row=1,column=4).value='PC_ASPECTS'
            sheet.cell(row=1,column=5).value='Age'
            sheet.cell(row=1,column=6).value='pre_BUN'
            sheet.cell(row=1,column=7).value='procedural time'   
            csv_file_name = 'for_predictions.csv'
            csv_exporter.save(csv_file_name)
            if os.path.exists(csv_file_name):
                data = open(csv_file_name, 'rb').read()
                b64 = base64.b64encode(data).decode('UTF-8')
                href = f'<a href="data:file/csv;base64,{b64}" download="{csv_file_name}">Download csv file</a>'
                st.markdown(href, unsafe_allow_html=True)
            else:
                 st.error(f"文件 '{csv_file_name}' 找不到，请检查文件生成的过程。")
            csv_exporter.close()

            file_upload = st.file_uploader("Upload CSV file for predictions", type=["csv"])

            if file_upload is not None: 
                try: 
                    data = pd.read_csv(file_upload, sep=',', error_bad_lines=False) 

                    if 'MRSI' in data.columns: 
                        y_true = data['MRSI'].values 
                        predictions = model4.predict_proba(data)[:, 1] 
                        predictions_df = pd.DataFrame(predictions, columns=['Predictions']) 
                        st.write(predictions)
                        result_data = data.copy() 
                        result_data['Predictions'] = predictions_df 
                        result_file_path = 'predictions_with_results.csv' 
                        result_data.to_csv(result_file_path, index=False) 
 
                        with open(result_file_path, 'rb') as f: 
                            output_data = f.read()
                            b64 = base64.b64encode(output_data).decode('UTF-8')
                            download_link = f'<a href="data:file/csv;base64,{b64}" download="predictions_with_results.csv">Download predictions with results</a>'
                            st.markdown(download_link, unsafe_allow_html=True)

                        add_data = st.selectbox('Outcome for Learning', [0, 1])
                
                        if st.button('Add Data for Learning'): 
                            X = data.drop(columns=['MRSI']) 
                            y = data['MRSI'] 
                            rf_model4 = model4.named_steps['trained_model']
                            intra_weighted_forest2 = DynamicWeightedForest(rf_model4.estimators_)
                            new_tree = DecisionTreeClassifier(random_state=42)
                            new_tree.fit(X, y) 
                            intra_weighted_forest2.add_tree(new_tree)
                            intra_weighted_forest2.update_weights(X, y) 
                            st.success("New tree added and weights updated dynamically!")
                            
                        def plot_combined_graphs(y_true, y_scores):
                            fig, axs = plt.subplots(1, 2, figsize=(14, 6))
                            fpr, tpr, _ = roc_curve(y_true, y_scores)
                            roc_auc = auc(fpr, tpr)
                            axs[0].plot(fpr, tpr, color='blue', label='ROC curve (area = %0.2f)' % roc_auc)
                            axs[0].plot([0, 1], [0, 1], color='red', linestyle='--')
                            axs[0].set_xlim([0.0, 1.0])
                            axs[0].set_ylim([0.0, 1.05])
                            axs[0].set_xlabel('False Positive Rate')
                            axs[0].set_ylabel('True Positive Rate')
                            axs[0].set_title('Receiver Operating Characteristic (ROC)')
                            axs[0].legend(loc='lower right')
                            axs[0].grid()

                            prob_true, prob_pred = calibration_curve(y_true, y_scores, n_bins=10)
                            axs[1].plot(prob_pred, prob_true, marker='o', label='Calibrated Model', color='b')
                            axs[1].plot([0, 1], [0, 1], linestyle='--', label='Perfectly Calibrated', color='r')
                            axs[1].set_title('Brier Score Calibration Plot')
                            axs[1].set_xlabel('Mean Predicted Probability')
                            axs[1].set_xlim([0, 1])
                            axs[1].set_ylim([0, 1])
                            axs[1].legend()
                            axs[1].grid()
                            st.pyplot(fig)

                        if len(data) >= 10:
                            y_pred = (predictions > 0.5).astype(int)
                            accuracy = accuracy_score(y_true, y_pred)
                            recall = recall_score(y_true, y_pred)
                            precision = precision_score(y_true, y_pred)
                            f1 = f1_score(y_true, y_pred)
                            roc_auc = auc(*roc_curve(y_true, predictions)[:2])

                            st.write(f"Accuracy: {accuracy:.2f}")
                            st.write(f"Recall: {recall:.2f}")
                            st.write(f"Precision: {precision:.2f}")
                            st.write(f"F1 Score: {f1:.2f}")
                            st.write(f"AUC: {roc_auc:.2f}")
                            brier_score = brier_score_loss(y_true, predictions)
                            st.write(f"Brier Score: {brier_score:.2f}")

                            plot_combined_graphs(y_true, predictions)
                        else:
                            st.warning("Not enough samples for ROC curve plotting. Please upload at least 10 samples.") 

                            st.write(predictions_df) 
                    else:                      
                        predictions = model4.predict_proba(data)[:,1] 
                        predictions = pd.DataFrame(predictions,columns = ['Predictions'])
                        st.write(predictions)
                        result_data = data.copy() 
                        result_data['Predictions'] = predictions 
                        result_file_path = 'predictions_with_results.csv'
                        result_data.to_csv(result_file_path, index=False)
                        with open(result_file_path, 'rb') as f:
                            data = f.read()
                            b64 = base64.b64encode(data).decode('UTF-8')
                            download_link = f'<a href="data:file/csv;base64,{b64}" download="predictions_with_results.csv">Download predictions with results</a>'
                            st.markdown(download_link, unsafe_allow_html=True)

                except Exception as e: 
                    st.error(f"Error reading the CSV file: {e}") 
            

        elif prediction_type == "Postoperative_number":
            st.subheader("Postoperative Number Prediction")
            st.write("This section will handle postoperative number predictions.please fill in the blanks with corresponding data. After that,click on the Predict button at the bottom to see the prediction of the classifier.")
            Age = st.number_input('Age', min_value = 0, max_value = 120, value = 60)
            GCS= st.number_input('GCS', min_value = 0,max_value = 15 ,value = 10)  
            PC_ASPECTS = st.number_input('PC_ASPECTS', min_value = 0.0,max_value = 10.0,value = 8.0)
            procedural_time = st.number_input('procedural time', min_value=0.00, max_value=350.00, value=60.00)
            post_eGFR = st.number_input('post_eGFR', min_value = 10.00,max_value = 250.00,value = 111.5) 
            post_NIHSS = st.number_input('post_NIHSS', min_value = 4,max_value = 38,value = 10) 
                            
            output=""
            features = { 
                'Age': Age, 
                'GCS': GCS, 
                'PC_ASPECTS': PC_ASPECTS,
                'procedural time': procedural_time, 
                'post_eGFR': post_eGFR,
                'post_NIHSS': post_NIHSS            
                
                  }

            print(features) 

            input_df = pd.DataFrame([features])

            print(input_df) 
            
            post_weighted_forest = DynamicWeightedForest(model5.estimators_)
    
            if st.button('Predict'): 
                output = model5.predict_proba(input_df)[:,1] 
                explainer = shap.Explainer(model5)
                shap_values = explainer.shap_values(input_df)
                st.write(' Based on feature values, predicted possibility of good functional outcome is '+ str(output))
                st_shap(shap.force_plot(explainer.expected_value[1], shap_values[1],input_df))
                shap_df = pd.DataFrame({
                    'Feature': input_df.columns,
                    'SHAP Value': shap_values[1].flatten() 
                 })
                st.write("SHAP values for each feature:")
                st.dataframe(shap_df)
        
            label = st.selectbox('Outcome for Learning', [0, 1])
            if st.button('Add Data for Learning'): 
                new_tree = DecisionTreeClassifier(random_state=42)
                new_tree.fit(input_df, [label])
                post_weighted_forest.add_tree(new_tree)
                post_weighted_forest.update_weights(input_df, [label])
                st.success("New tree added and weights updated dynamically!")

        elif prediction_type == "Postoperative_batch":
            st.subheader("Postoperative Batch Prediction")
            st.write("This section will handle postoperative batch predictions.Please click on the link to download the form and fill in the corresponding data. After that, click on the Browse files button to upload file for prediciton, you can see the prediction of the classifier at the bottom. This page supports batch prediction of the outcome of multiple patients at one time, and can predict the outcome of patients with missing values.")
            def plot_roc_curve(y_true, y_scores): 
                fpr, tpr, thresholds = roc_curve(y_true, y_scores) 
                roc_auc = auc(fpr, tpr)

                plt.figure(figsize=(10, 6)) 
                plt.plot(fpr, tpr, color='blue', label='ROC curve (area = %0.2f)' % roc_auc) 
                plt.plot([0, 1], [0, 1], color='red', linestyle='--') 
                plt.xlim([0.0, 1.0]) 
                plt.ylim([0.0, 1.05]) 
                plt.xlabel('False Positive Rate') 
                plt.ylabel('True Positive Rate')
                plt.title('Receiver Operating Characteristic (ROC)') 
                plt.legend(loc='lower right') 
                plt.grid() 
                st.pyplot(plt) 

            csv_exporter = openpyxl.Workbook()
            sheet = csv_exporter.active
            sheet.cell(row=1, column=1).value = 'Age'
            sheet.cell(row=1, column=2).value = 'GCS' 
            sheet.cell(row=1, column=3).value = 'PC_ASPECTS' 
            sheet.cell(row=1, column=4).value = 'procedural time' 
            sheet.cell(row=1, column=5).value = 'post_eGFR' 
            sheet.cell(row=1, column=6).value = 'post_NIHSS'     
            csv_file_name = 'for_predictions.csv'
            csv_exporter.save(csv_file_name)
            if os.path.exists(csv_file_name):
                data = open(csv_file_name, 'rb').read()
                b64 = base64.b64encode(data).decode('UTF-8')
                href = f'<a href="data:file/csv;base64,{b64}" download="{csv_file_name}">Download csv file</a>'
                st.markdown(href, unsafe_allow_html=True)
            else:
                 st.error(f"文件 '{csv_file_name}' 找不到，请检查文件生成的过程。")
            csv_exporter.close()

            file_upload = st.file_uploader("Upload CSV file for predictions", type=["csv"])

            if file_upload is not None: 
                try: 
                    data = pd.read_csv(file_upload, sep=',', error_bad_lines=False) 

                    if 'MRSI' in data.columns: 
                        y_true = data['MRSI'].values 
                        predictions = model6.predict_proba(data)[:, 1] 
                        predictions_df = pd.DataFrame(predictions, columns=['Predictions']) 
                        st.write(predictions)
                        result_data = data.copy() 
                        result_data['Predictions'] = predictions_df 
                        result_file_path = 'predictions_with_results.csv' 
                        result_data.to_csv(result_file_path, index=False) 
 
                        with open(result_file_path, 'rb') as f: 
                            output_data = f.read()
                            b64 = base64.b64encode(output_data).decode('UTF-8')
                            download_link = f'<a href="data:file/csv;base64,{b64}" download="predictions_with_results.csv">Download predictions with results</a>'
                            st.markdown(download_link, unsafe_allow_html=True)

                        add_data = st.selectbox('Outcome for Learning', [0, 1])
                
                        if st.button('Add Data for Learning'): 
                            X = data.drop(columns=['MRSI']) 
                            y = data['MRSI'] 
                            rf_model6 = model6.named_steps['trained_model']
                            post_weighted_forest2 = DynamicWeightedForest(rf_model6.estimators_)
                            new_tree = DecisionTreeClassifier(random_state=42)
                            new_tree.fit(X, y) 
                            post_weighted_forest2.add_tree(new_tree)
                            post_weighted_forest2.update_weights(X, y) 
                            st.success("New tree added and weights updated dynamically!")

                        def plot_combined_graphs(y_true, y_scores):
                            fig, axs = plt.subplots(1, 2, figsize=(14, 6))
                            fpr, tpr, _ = roc_curve(y_true, y_scores)
                            roc_auc = auc(fpr, tpr)
                            axs[0].plot(fpr, tpr, color='blue', label='ROC curve (area = %0.2f)' % roc_auc)
                            axs[0].plot([0, 1], [0, 1], color='red', linestyle='--')
                            axs[0].set_xlim([0.0, 1.0])
                            axs[0].set_ylim([0.0, 1.05])
                            axs[0].set_xlabel('False Positive Rate')
                            axs[0].set_ylabel('True Positive Rate')
                            axs[0].set_title('Receiver Operating Characteristic (ROC)')
                            axs[0].legend(loc='lower right')
                            axs[0].grid()

                            prob_true, prob_pred = calibration_curve(y_true, y_scores, n_bins=10)
                            axs[1].plot(prob_pred, prob_true, marker='o', label='Calibrated Model', color='b')
                            axs[1].plot([0, 1], [0, 1], linestyle='--', label='Perfectly Calibrated', color='r')
                            axs[1].set_title('Brier Score Calibration Plot')
                            axs[1].set_xlabel('Mean Predicted Probability')
                            axs[1].set_ylabel('Fraction of Positives')
                            axs[1].set_xlim([0, 1])
                            axs[1].set_ylim([0, 1])
                            axs[1].legend()
                            axs[1].grid()
                            st.pyplot(fig)

                        if len(data) >= 10:
                            y_pred = (predictions > 0.5).astype(int)
                            accuracy = accuracy_score(y_true, y_pred)
                            recall = recall_score(y_true, y_pred)
                            precision = precision_score(y_true, y_pred)
                            f1 = f1_score(y_true, y_pred)
                            roc_auc = auc(*roc_curve(y_true, predictions)[:2])

                            st.write(f"Accuracy: {accuracy:.2f}")
                            st.write(f"Recall: {recall:.2f}")
                            st.write(f"Precision: {precision:.2f}")
                            st.write(f"F1 Score: {f1:.2f}")
                            st.write(f"AUC: {roc_auc:.2f}")
                            brier_score = brier_score_loss(y_true, predictions)
                            st.write(f"Brier Score: {brier_score:.2f}")
                            plot_combined_graphs(y_true, predictions)

                            if roc_auc < 0.75:
                                st.write("AUC is below 0.75. Retraining with Class Incremental Random Forests.")
                    
                                X = data.drop(columns=['MRSI']) 
                                y = data['MRSI'] 
                                cifr_model.partial_fit(X, y)  
                        else:
                            st.warning("Not enough samples for ROC curve plotting. Please upload at least 10 samples.") 

                            st.write(predictions_df) 
                    else:                      
                        predictions = model6.predict_proba(data)[:,1] 
                        predictions = pd.DataFrame(predictions,columns = ['Predictions'])
                        st.write(predictions)
                        result_data = data.copy() 
                        result_data['Predictions'] = predictions 
                        result_file_path = 'predictions_with_results.csv'
                        result_data.to_csv(result_file_path, index=False)
                        with open(result_file_path, 'rb') as f:
                            data = f.read()
                            b64 = base64.b64encode(data).decode('UTF-8')
                            download_link = f'<a href="data:file/csv;base64,{b64}" download="predictions_with_results.csv">Download predictions with results</a>'
                            st.markdown(download_link, unsafe_allow_html=True)

                except Exception as e: 
                    st.error(f"Error reading the CSV file: {e}")

    else:  # Other Features
        st.title("Other Features")
        st.write("Here you can describe other features of your application.")
  
def main():
    initialize_database()  

    if 'is_logged_in' not in st.session_state:
        st.session_state['is_logged_in'] = False

    if st.session_state['is_logged_in']:
        prediction_page()  
    else:
        login_or_register = st.sidebar.selectbox("Select an action:", ("Login", "Register"), key="login_register_selectbox")
        if login_or_register == "Login":
            login_page()
        else:
            register_page()

if __name__ == "__main__":
    main()
