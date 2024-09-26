#!/usr/bin/env python
# coding: utf-8

# In[1]:


import streamlit as st
import pandas as pd
import numpy as np
import pickle
import shap
shap.initjs()
from sklearn.preprocessing import OneHotEncoder
from PIL import Image
from pycaret import classification
from pycaret.classification import setup, get_logs,create_model, predict_model, tune_model, save_model,load_model


# In[2]:


import streamlit.components.v1 as components

st.markdown(
    f"""
    <style>
        .appview-container .main .block-container{{
            max-width: {1150}px;
            padding-top: {5}rem;
            padding-right: {10}rem;
            padding-left: {10}rem;
            padding-bottom: {5}rem;
        }}
        .reportview-container .main {{
            color: white;
            background-color: black;
        }}
    </style>
    """
    ,unsafe_allow_html=True,
)


# In[5]:

# In[4]:
image = Image.open('hospital1.webp')
st.image(image, use_column_width=True)

import joblib
model = pickle.load(open( "model1.p", "rb" ))
model2 = load_model('tuned_rf_pre_model')
model3 = joblib.load('tuned_rf_peri.pkl')
model4 = load_model('tuned_rf_peri_model')
model5 = joblib.load('tuned_rf43.pkl')
model6 = load_model('tuned_rf43_model')

def st_shap(plot, height=None):
    shap_html = f"<head>{shap.getjs()}</head><body>{plot.html()}</body>"
    components.html(shap_html, height=height)


# In[9]:


def run():

    st.title('Functional outcome prediction  App for patients with  anterior circulation large vessel occlusion after mechanical thrombectomy')

add_selectbox = st.sidebar.selectbox(
    "How would you like to predict?",
    ("Preoperative_number","Preoperative_batch", "Perioperative_number","Perioperative_batch", "Postoperative_number","Postoperative_batch"))


# In[11]:


if add_selectbox == 'Preoperative_number': 
    st.write('This is a web app to predict 90-day outcome of acute ischaemic stroke patients with mechanical thrombectomy based on several features that you can see in the sidebar. Please adjust the value of each feature. After that, click on the Predict button at the bottom to see the prediction of the classifier.')
    
    NIHSS = st.number_input('NIHSS', min_value = 4,max_value = 38,value = 10) 
    GCS= st.number_input('GCS', min_value = 0,max_value = 15 ,value = 10) 
    pre_eGFR = st.number_input('pre_eGFR', min_value = 10.00,max_value = 250.00,value = 111.5)
    pre_glucose = st.number_input('pre_glucose', min_value = 2.50, max_value = 25.00, value = 7.78)
    PC_ASPECT = st.number_input('PC_ASPECT', min_value = 0.0,max_value = 10.0,value = 8.0)
    pre_fibrinogen = st.number_input('pre_fibrinogen', min_value = 0.20,max_value = 50.00,value = 3.20)
    Gender = st.selectbox('Gender', [1, 0])
    output=""

    features = { 
        'NIHSS': NIHSS, 
        'GCS': GCS, 
        'pre_eGFR': pre_eGFR,
        'pre_glucose': pre_glucose, 
        'PC_ASPECT': PC_ASPECT, 
        'pre_fibrinogen': pre_fibrinogen, 
        'Gender': 1 - Gender 
          }

    print(features) 


    input_df = pd.DataFrame([features]) 
    print(input_df) 
    
    if st.button('Predict'): 
        output = model.predict_proba(input_df)[:,0] 
        explainer = shap.Explainer(model)
        shap_values = explainer.shap_values(input_df)
        st.write(' Based on feature values, predicted possibility of good functional outcome is '+ str(output))
        st_shap(shap.force_plot(explainer.expected_value[0], shap_values[0],input_df))
        shap_df = pd.DataFrame({
            'Feature': input_df.columns,
            'SHAP Value': shap_values[1].flatten() 
         })

        st.write("SHAP values for each feature:")
        st.dataframe(shap_df)


import streamlit as st 
import pandas as pd 
import openpyxl 
import base64 
import matplotlib.pyplot as plt
from sklearn.calibration import calibration_curve
from sklearn.metrics import brier_score_loss, accuracy_score, recall_score, precision_score, f1_score, roc_auc_score, roc_curve
from sklearn.metrics import roc_curve, auc, accuracy_score, recall_score, precision_score, f1_score 

def plot_roc_curve(y_true, y_scores): 
    fpr, tpr, thresholds = roc_curve(y_true, y_scores) 
    roc_auc = auc(fpr, tpr)

    plt.figure(figsize=(10, 6)) 
    plt.plot(fpr, tpr, color='blue', label='ROC curve (area = %0.2f)' % roc_auc) 
    plt.plot([0, 1], [0, 1], color='red', linestyle='--') 
    plt.xlim([0.0, 1.0]) 
    plt.ylim([0.0, 1.05]) 
    plt.xlabel('False Positive Rate') 
    plt.ylabel('True Positive Rate')
    plt.title('Receiver Operating Characteristic (ROC)') 
    plt.legend(loc='lower right') 
    plt.grid() 
    st.pyplot(plt) 

if add_selectbox == 'Preoperative_batch': 
    st.write('This is a web app to predict 90-day outcome of acute ischaemic stroke patients with mechanical thrombectomy based on several features. Please click on the link to download the form and fill in the corresponding data. After that, click on the Browse files button to upload file for prediciton, you can see the prediction of the classifier at the bottom. This page supports batch prediction of the outcome of multiple patients at one time, and can predict the outcome of patients with missing values.') 

    csv_exporter=openpyxl.Workbook()
    sheet=csv_exporter.active
    sheet.cell(row=1,column=1).value='NIHSS'
    sheet.cell(row=1,column=2).value='GCS'
    sheet.cell(row=1,column=3).value='pre_eGFR'
    sheet.cell(row=1,column=4).value='pre_glucose'
    sheet.cell(row=1,column=5).value='PC_ASPECT'
    sheet.cell(row=1,column=6).value='pre_fibrinogen'
    sheet.cell(row=1,column=7).value='Gender'
    csv_exporter.save('for predictions.csv')#注意！文件此时保存在内存中且为字节格式文件

    data = open('for_predictions.csv', 'rb').read() 
    b64 = base64.b64encode(data).decode('UTF-8') 
    href = f'<a href="data:file/data;base64,{b64}" download="for_predictions.csv">Download csv file</a>'
    st.markdown(href, unsafe_allow_html=True) 
    csv_exporter.close() 

    file_upload = st.file_uploader("Upload CSV file for predictions", type=["csv"]) 

    if file_upload is not None: 
        try: 
            data = pd.read_csv(file_upload, sep=',', error_bad_lines=False) 

            if 'MRSI' in data.columns: 
                y_true = data['MRSI'].values 
                predictions = model2.predict_proba(data)[:, 1] 
                predictions_df = pd.DataFrame(predictions, columns=['Predictions']) 
                st.write(predictions)
                result_data = data.copy() 
                result_data['Predictions'] = predictions_df 
                result_file_path = 'predictions_with_results.csv' 
                result_data.to_csv(result_file_path, index=False) 
 
                with open(result_file_path, 'rb') as f: 
                    output_data = f.read()
                    b64 = base64.b64encode(output_data).decode('UTF-8')
                    download_link = f'<a href="data:file/csv;base64,{b64}" download="predictions_with_results.csv">Download predictions with results</a>'
                    st.markdown(download_link, unsafe_allow_html=True)

                add_data = st.selectbox('Outcome for Learning', [0, 1])
                if st.button('Add Data for Learning'): 
                    X = data.drop(columns=['MRSI']) 
                    y = data['MRSI'] 
                    model2.fit(X, y) 
                    st.success("New data has been added to the model for continuous learning!")

                def plot_combined_graphs(y_true, y_scores):
                    fig, axs = plt.subplots(1, 2, figsize=(14, 6))
                    fpr, tpr, _ = roc_curve(y_true, y_scores)
                    roc_auc = auc(fpr, tpr)
                    axs[0].plot(fpr, tpr, color='blue', label='ROC curve (area = %0.2f)' % roc_auc)
                    axs[0].plot([0, 1], [0, 1], color='red', linestyle='--')
                    axs[0].set_xlim([0.0, 1.0])
                    axs[0].set_ylim([0.0, 1.05])
                    axs[0].set_xlabel('False Positive Rate')
                    axs[0].set_ylabel('True Positive Rate')
                    axs[0].set_title('Receiver Operating Characteristic (ROC)')
                    axs[0].legend(loc='lower right')
                    axs[0].grid()

# 计算并绘制 Brier 校正图
                    prob_true, prob_pred = calibration_curve(y_true, y_scores, n_bins=10)
                    axs[1].plot(prob_pred, prob_true, marker='o', label='Calibrated Model', color='b')
                    axs[1].plot([0, 1], [0, 1], linestyle='--', label='Perfectly Calibrated', color='r')
                    axs[1].set_title('Brier Score Calibration Plot')
                    axs[1].set_xlabel('Mean Predicted Probability')
                    axs[1].set_ylabel('Fraction of Positives')
                    axs[1].set_xlim([0, 1])
                    axs[1].set_ylim([0, 1])
                    axs[1].legend()
                    axs[1].grid()
                    st.pyplot(fig)

# 在您的主代码中用于展示 ROC 和 Brier 校正图
                if len(data) >= 10:
                    y_pred = (predictions > 0.5).astype(int)
                    accuracy = accuracy_score(y_true, y_pred)
                    recall = recall_score(y_true, y_pred)
                    precision = precision_score(y_true, y_pred)
                    f1 = f1_score(y_true, y_pred)
                    roc_auc = auc(*roc_curve(y_true, predictions)[:2])

                    st.write(f"Accuracy: {accuracy:.2f}")
                    st.write(f"Recall: {recall:.2f}")
                    st.write(f"Precision: {precision:.2f}")
                    st.write(f"F1 Score: {f1:.2f}")
                    st.write(f"AUC: {roc_auc:.2f}")
                    brier_score = brier_score_loss(y_true, predictions)
                    st.write(f"Brier Score: {brier_score:.2f}")

# 绘制并展示ROC曲线和Brier校正图
                    plot_combined_graphs(y_true, predictions)

                else:
                    st.warning("Not enough samples for ROC curve plotting. Please upload at least 10 samples.") 

                    st.write(predictions_df) 
            else:                      
                predictions = model2.predict_proba(data)[:,1] 
                predictions = pd.DataFrame(predictions,columns = ['Predictions'])
                st.write(predictions)
                result_data = data.copy() 
                result_data['Predictions'] = predictions 
                result_file_path = 'predictions_with_results.csv'
                result_data.to_csv(result_file_path, index=False)
                with open(result_file_path, 'rb') as f:
                    data = f.read()
                    b64 = base64.b64encode(data).decode('UTF-8')
                    download_link = f'<a href="data:file/csv;base64,{b64}" download="predictions_with_results.csv">Download predictions with results</a>'
                    st.markdown(download_link, unsafe_allow_html=True)

        except Exception as e: 
            st.error(f"Error reading the CSV file: {e}") 

        
        
if add_selectbox == 'Perioperative_number': 
    st.write('This is a web app to predict 90-day outcome of acute ischaemic stroke patients with mechanical thrombectomy based on several features that you can see in the sidebar. Please adjust the value of each feature. After that, click on the Predict button at the bottom to see the prediction of the classifier.')

    NIHSS = st.number_input('NIHSS', min_value = 4,max_value = 38,value = 10) 
    GCS= st.number_input('GCS', min_value = 0,max_value = 15 ,value = 10) 
    pre_eGFR = st.number_input('pre_eGFR', min_value = 10.00,max_value = 250.00,value = 111.5)
    pre_glucose = st.number_input('pre_glucose', min_value = 2.50, max_value = 25.00, value = 7.78)
    PC_ASPECT = st.number_input('PC_ASPECT', min_value = 0.0,max_value = 10.0,value = 8.0)
    Duration = st.number_input('Duration', min_value = 0.00,max_value = 350.00,value = 60.00)
    pre_fibrinogen = st.number_input('pre_fibrinogen', min_value = 0.20,max_value = 50.00,value = 3.20)
    Gender = st.selectbox('Female', [1, 0])
    output=""

    features = { 
        'NIHSS': NIHSS, 
        'GCS': GCS, 
        'pre_eGFR': pre_eGFR,
        'pre_glucose': pre_glucose, 
        'PC_ASPECT': PC_ASPECT, 
        'Duration': Duration, 
        'pre_fibrinogen': pre_fibrinogen, 
        'Female': Gender
          }


    print(features) 

    input_df = pd.DataFrame([features])

    print(input_df) 


    
    if st.button('Predict'): 
        output = model3.predict_proba(input_df)[:,1] 
        explainer = shap.Explainer(model3)
        shap_values = explainer.shap_values(input_df)
        st.write(' Based on feature values, predicted possibility of good functional outcome is '+ str(output))
        st_shap(shap.force_plot(explainer.expected_value[1], shap_values[1],input_df))
        shap_df = pd.DataFrame({
            'Feature': input_df.columns,
            'SHAP Value': shap_values[1].flatten() 
         })
        st.write("SHAP values for each feature:")
        st.dataframe(shap_df)
        
    label = st.selectbox('Outcome for Learning', [0, 1])
    if st.button('Add Data for Learning'): 
        model3.fit(input_df, [label])
        st.success("New data has been added to the model for continuous learning!")
        
import streamlit as st 
import pandas as pd 
import openpyxl 
import base64 
import matplotlib.pyplot as plt
from sklearn.metrics import roc_curve, auc, accuracy_score, recall_score, precision_score, f1_score 

def plot_roc_curve(y_true, y_scores): 
    fpr, tpr, thresholds = roc_curve(y_true, y_scores) 
    roc_auc = auc(fpr, tpr)

    plt.figure(figsize=(10, 6)) 
    plt.plot(fpr, tpr, color='blue', label='ROC curve (area = %0.2f)' % roc_auc) 
    plt.plot([0, 1], [0, 1], color='red', linestyle='--') 
    plt.xlim([0.0, 1.0]) 
    plt.ylim([0.0, 1.05]) 
    plt.xlabel('False Positive Rate') 
    plt.ylabel('True Positive Rate')
    plt.title('Receiver Operating Characteristic (ROC)') 
    plt.legend(loc='lower right') 
    plt.grid() 
    st.pyplot(plt) 

if add_selectbox == 'Perioperative_batch': 
    st.write('This is a web app to predict 90-day outcome of acute ischaemic stroke patients with mechanical thrombectomy based on several features. Please click on the link to download the form and fill in the corresponding data. After that, click on the Browse files button to upload file for prediciton, you can see the prediction of the classifier at the bottom. This page supports batch prediction of the outcome of multiple patients at one time, and can predict the outcome of patients with missing values.') 

    csv_exporter=openpyxl.Workbook()
    sheet=csv_exporter.active
    sheet.cell(row=1,column=1).value='NIHSS'
    sheet.cell(row=1,column=2).value='GCS'
    sheet.cell(row=1,column=3).value='pre_eGFR'
    sheet.cell(row=1,column=4).value='pre_glucose'
    sheet.cell(row=1,column=5).value='PC_ASPECT'
    sheet.cell(row=1,column=6).value='Duration'
    sheet.cell(row=1,column=7).value='pre_fibrinogen'
    sheet.cell(row=1,column=8).value='Gender'
    csv_exporter.save('for predictions.csv')

    data = open('for_predictions.csv', 'rb').read() 
    b64 = base64.b64encode(data).decode('UTF-8') 
    href = f'<a href="data:file/data;base64,{b64}" download="for_predictions.csv">Download csv file</a>'
    st.markdown(href, unsafe_allow_html=True) 
    csv_exporter.close() 

    file_upload = st.file_uploader("Upload CSV file for predictions", type=["csv"]) 

    if file_upload is not None: 
        try: 
            data = pd.read_csv(file_upload, sep=',', error_bad_lines=False) 

            if 'MRSI' in data.columns: 
                y_true = data['MRSI'].values 
                predictions = model4.predict_proba(data)[:, 1] 
                predictions_df = pd.DataFrame(predictions, columns=['Predictions']) 
                st.write(predictions)
                result_data = data.copy() 
                result_data['Predictions'] = predictions_df 
                result_file_path = 'predictions_with_results.csv' 
                result_data.to_csv(result_file_path, index=False) 
 
                with open(result_file_path, 'rb') as f: 
                    output_data = f.read()
                    b64 = base64.b64encode(output_data).decode('UTF-8')
                    download_link = f'<a href="data:file/csv;base64,{b64}" download="predictions_with_results.csv">Download predictions with results</a>'
                    st.markdown(download_link, unsafe_allow_html=True)

                add_data = st.selectbox('Outcome for Learning', [0, 1])
                
                if st.button('Add Data for Learning'): 
                    X = data.drop(columns=['MRSI']) 
                    y = data['MRSI'] 
                    model4.fit(X, y) 
                    st.success("New data has been added to the model for continuous learning!")

                def plot_combined_graphs(y_true, y_scores):
                    fig, axs = plt.subplots(1, 2, figsize=(14, 6))
                    fpr, tpr, _ = roc_curve(y_true, y_scores)
                    roc_auc = auc(fpr, tpr)
                    axs[0].plot(fpr, tpr, color='blue', label='ROC curve (area = %0.2f)' % roc_auc)
                    axs[0].plot([0, 1], [0, 1], color='red', linestyle='--')
                    axs[0].set_xlim([0.0, 1.0])
                    axs[0].set_ylim([0.0, 1.05])
                    axs[0].set_xlabel('False Positive Rate')
                    axs[0].set_ylabel('True Positive Rate')
                    axs[0].set_title('Receiver Operating Characteristic (ROC)')
                    axs[0].legend(loc='lower right')
                    axs[0].grid()

# 计算并绘制 Brier 校正图
                    prob_true, prob_pred = calibration_curve(y_true, y_scores, n_bins=10)
                    axs[1].plot(prob_pred, prob_true, marker='o', label='Calibrated Model', color='b')
                    axs[1].plot([0, 1], [0, 1], linestyle='--', label='Perfectly Calibrated', color='r')
                    axs[1].set_title('Brier Score Calibration Plot')
                    axs[1].set_xlabel('Mean Predicted Probability')
                    axs[1].set_ylabel('Fraction of Positives')
                    axs[1].set_xlim([0, 1])
                    axs[1].set_ylim([0, 1])
                    axs[1].legend()
                    axs[1].grid()
                    st.pyplot(fig)

# 在您的主代码中用于展示 ROC 和 Brier 校正图
                if len(data) >= 10:
                    y_pred = (predictions > 0.5).astype(int)
                    accuracy = accuracy_score(y_true, y_pred)
                    recall = recall_score(y_true, y_pred)
                    precision = precision_score(y_true, y_pred)
                    f1 = f1_score(y_true, y_pred)
                    roc_auc = auc(*roc_curve(y_true, predictions)[:2])

                    st.write(f"Accuracy: {accuracy:.2f}")
                    st.write(f"Recall: {recall:.2f}")
                    st.write(f"Precision: {precision:.2f}")
                    st.write(f"F1 Score: {f1:.2f}")
                    st.write(f"AUC: {roc_auc:.2f}")
                    brier_score = brier_score_loss(y_true, predictions)
                    st.write(f"Brier Score: {brier_score:.2f}")

# 绘制并展示ROC曲线和Brier校正图
                    plot_combined_graphs(y_true, predictions)
                else:
                    st.warning("Not enough samples for ROC curve plotting. Please upload at least 10 samples.") 

                    st.write(predictions_df) 
            else:                      
                predictions = model4.predict_proba(data)[:,1] 
                predictions = pd.DataFrame(predictions,columns = ['Predictions'])
                st.write(predictions)
                result_data = data.copy() 
                result_data['Predictions'] = predictions 
                result_file_path = 'predictions_with_results.csv'
                result_data.to_csv(result_file_path, index=False)
                with open(result_file_path, 'rb') as f:
                    data = f.read()
                    b64 = base64.b64encode(data).decode('UTF-8')
                    download_link = f'<a href="data:file/csv;base64,{b64}" download="predictions_with_results.csv">Download predictions with results</a>'
                    st.markdown(download_link, unsafe_allow_html=True)

        except Exception as e: 
            st.error(f"Error reading the CSV file: {e}") 
            
            
        
if add_selectbox == 'Postoperative_number': 
    st.write('This is a web app to predict 90-day outcome of acute ischaemic stroke patients with mechanical thrombectomy based on several features that you can see in the sidebar. Please adjust the value of each feature. After that, click on the Predict button at the bottom to see the prediction of the classifier.')
    
    pre_glucose = st.number_input('pre_glucose', min_value = 2.50, max_value = 25.00, value = 7.78)
    PC_ASPECT = st.number_input('PC_ASPECT', min_value = 0.0,max_value = 10.0,value = 8.0)
    post_eGFR = st.number_input('post_eGFR', min_value = 10.00,max_value = 250.00,value = 111.5) 
    post_NIHSS = st.number_input('post_NIHSS', min_value = 4,max_value = 38,value = 10) 
    GCS= st.number_input('GCS', min_value = 0,max_value = 15 ,value = 10)      
    Duration = st.number_input('Duration', min_value = 0.00,max_value = 350.00,value = 60.00)
    output=""

    features = { 
        'pre_glucose': pre_glucose, 
        'PC_ASPECT': PC_ASPECT, 
        'post_eGFR': post_eGFR,
        'post_NIHSS': post_NIHSS, 
        'GCS': GCS, 
        'Duration': Duration, 
          }


    print(features) 

    input_df = pd.DataFrame([features])

    print(input_df) 


    
    if st.button('Predict'): 
        output = model5.predict_proba(input_df)[:,1] 
        explainer = shap.Explainer(model5)
        shap_values = explainer.shap_values(input_df)
        st.write(' Based on feature values, predicted possibility of good functional outcome is '+ str(output))
        st_shap(shap.force_plot(explainer.expected_value[1], shap_values[1],input_df))
        shap_df = pd.DataFrame({
            'Feature': input_df.columns,
            'SHAP Value': shap_values[1].flatten() 
         })
        st.write("SHAP values for each feature:")
        st.dataframe(shap_df)
        
    label = st.selectbox('Outcome for Learning', [0, 1])
    if st.button('Add Data for Learning'): 
        model5.fit(input_df, [label])
        st.success("New data has been added to the model for continuous learning!")

import streamlit as st 
import pandas as pd 
import openpyxl 
import base64 
import matplotlib.pyplot as plt
from sklearn.metrics import roc_curve, auc, accuracy_score, recall_score, precision_score, f1_score 



def plot_roc_curve(y_true, y_scores): 
    fpr, tpr, thresholds = roc_curve(y_true, y_scores) 
    roc_auc = auc(fpr, tpr)

    plt.figure(figsize=(10, 6)) 
    plt.plot(fpr, tpr, color='blue', label='ROC curve (area = %0.2f)' % roc_auc) 
    plt.plot([0, 1], [0, 1], color='red', linestyle='--') 
    plt.xlim([0.0, 1.0]) 
    plt.ylim([0.0, 1.05]) 
    plt.xlabel('False Positive Rate') 
    plt.ylabel('True Positive Rate')
    plt.title('Receiver Operating Characteristic (ROC)') 
    plt.legend(loc='lower right') 
    plt.grid() 
    st.pyplot(plt) 

if add_selectbox == 'Postoperative_batch': 
    st.write('This is a web app to predict 90-day outcome of acute ischaemic stroke patients with mechanical thrombectomy based on several features. Please click on the link to download the form and fill in the corresponding data. After that, click on the Browse files button to upload file for prediciton, you can see the prediction of the classifier at the bottom. This page supports batch prediction of the outcome of multiple patients at one time, and can predict the outcome of patients with missing values.') 

    csv_exporter = openpyxl.Workbook() 
    sheet = csv_exporter.active 
    sheet.cell(row=1, column=1).value = 'pre_glucose'
    sheet.cell(row=1, column=2).value = 'PC_ASPECT' 
    sheet.cell(row=1, column=3).value = 'post_eGFR' 
    sheet.cell(row=1, column=4).value = 'post_NIHSS' 
    sheet.cell(row=1, column=5).value = 'GCS' 
    sheet.cell(row=1, column=6).value = 'Duration' 
    csv_exporter.save('for_predictions.csv') 

    data = open('for_predictions.csv', 'rb').read() 
    b64 = base64.b64encode(data).decode('UTF-8') 
    href = f'<a href="data:file/data;base64,{b64}" download="for_predictions.csv">Download csv file</a>'
    st.markdown(href, unsafe_allow_html=True) 
    csv_exporter.close() 

    file_upload = st.file_uploader("Upload CSV file for predictions", type=["csv"]) 

    if file_upload is not None: 
        try: 
            data = pd.read_csv(file_upload, sep=',', error_bad_lines=False) 

            if 'MRSI' in data.columns: 
                y_true = data['MRSI'].values 
                predictions = model6.predict_proba(data)[:, 1] 
                predictions_df = pd.DataFrame(predictions, columns=['Predictions']) 
                st.write(predictions)
                result_data = data.copy() 
                result_data['Predictions'] = predictions_df 
                result_file_path = 'predictions_with_results.csv' 
                result_data.to_csv(result_file_path, index=False) 
 
                with open(result_file_path, 'rb') as f: 
                    output_data = f.read()
                    b64 = base64.b64encode(output_data).decode('UTF-8')
                    download_link = f'<a href="data:file/csv;base64,{b64}" download="predictions_with_results.csv">Download predictions with results</a>'
                    st.markdown(download_link, unsafe_allow_html=True)

                add_data = st.selectbox('Outcome for Learning', [0, 1])
                
                if st.button('Add Data for Learning'): 
                    X = data.drop(columns=['MRSI']) 
                    y = data['MRSI'] 
                    model6.fit(X, y) 
                    st.success("New data has been added to the model for continuous learning!")

                def plot_combined_graphs(y_true, y_scores):
                    fig, axs = plt.subplots(1, 2, figsize=(14, 6))
                    fpr, tpr, _ = roc_curve(y_true, y_scores)
                    roc_auc = auc(fpr, tpr)
                    axs[0].plot(fpr, tpr, color='blue', label='ROC curve (area = %0.2f)' % roc_auc)
                    axs[0].plot([0, 1], [0, 1], color='red', linestyle='--')
                    axs[0].set_xlim([0.0, 1.0])
                    axs[0].set_ylim([0.0, 1.05])
                    axs[0].set_xlabel('False Positive Rate')
                    axs[0].set_ylabel('True Positive Rate')
                    axs[0].set_title('Receiver Operating Characteristic (ROC)')
                    axs[0].legend(loc='lower right')
                    axs[0].grid()

# 计算并绘制 Brier 校正图
                    prob_true, prob_pred = calibration_curve(y_true, y_scores, n_bins=10)
                    axs[1].plot(prob_pred, prob_true, marker='o', label='Calibrated Model', color='b')
                    axs[1].plot([0, 1], [0, 1], linestyle='--', label='Perfectly Calibrated', color='r')
                    axs[1].set_title('Brier Score Calibration Plot')
                    axs[1].set_xlabel('Mean Predicted Probability')
                    axs[1].set_ylabel('Fraction of Positives')
                    axs[1].set_xlim([0, 1])
                    axs[1].set_ylim([0, 1])
                    axs[1].legend()
                    axs[1].grid()
                    st.pyplot(fig)

# 在您的主代码中用于展示 ROC 和 Brier 校正图
                if len(data) >= 10:
                    y_pred = (predictions > 0.5).astype(int)
                    accuracy = accuracy_score(y_true, y_pred)
                    recall = recall_score(y_true, y_pred)
                    precision = precision_score(y_true, y_pred)
                    f1 = f1_score(y_true, y_pred)
                    roc_auc = auc(*roc_curve(y_true, predictions)[:2])

                    st.write(f"Accuracy: {accuracy:.2f}")
                    st.write(f"Recall: {recall:.2f}")
                    st.write(f"Precision: {precision:.2f}")
                    st.write(f"F1 Score: {f1:.2f}")
                    st.write(f"AUC: {roc_auc:.2f}")
                    brier_score = brier_score_loss(y_true, predictions)
                    st.write(f"Brier Score: {brier_score:.2f}")

# 绘制并展示ROC曲线和Brier校正图
                    plot_combined_graphs(y_true, predictions)

                else:
                    st.warning("Not enough samples for ROC curve plotting. Please upload at least 10 samples.") 

                    st.write(predictions_df) 
            else:                      
                predictions = model6.predict_proba(data)[:,1] 
                predictions = pd.DataFrame(predictions,columns = ['Predictions'])
                st.write(predictions)
                result_data = data.copy() 
                result_data['Predictions'] = predictions 
                result_file_path = 'predictions_with_results.csv'
                result_data.to_csv(result_file_path, index=False)
                with open(result_file_path, 'rb') as f:
                    data = f.read()
                    b64 = base64.b64encode(data).decode('UTF-8')
                    download_link = f'<a href="data:file/csv;base64,{b64}" download="predictions_with_results.csv">Download predictions with results</a>'
                    st.markdown(download_link, unsafe_allow_html=True)

        except Exception as e: 
            st.error(f"Error reading the CSV file: {e}") 



