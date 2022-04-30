#!/usr/bin/env python
# coding: utf-8

# In[16]:


import streamlit as st
import pandas as pd
import numpy as np
import pickle
from sklearn.preprocessing import OneHotEncoder
from PIL import Image
from xgboost import XGBClassifier

# In[17]:


model = pickle.load(open( "tuned_xgboostz4.p", "rb" ))


# In[18]:


def transform(data):    
        data['Collateral_status'] = pd.factorize(data['Collateral_status'])[0]
        return data


# In[19]:


def predict(model, input_df):
    transform_data = transform(input_df)
    predictions_df = model.predict(transform_data)
    predictions = predictions_df[0]
    return predictions_df


# In[20]:


def run():


 st.title('Functional outcome Classifier Web App')
st.write('This is a web app to predict your 90-day functional outcome based on several features. Please adjust the value of each feature. After that, click on the Predict button at the bottom to see the prediction of the classifier.')


# In[21]:


image = Image.open('hospital1.webp')
st.image(image, use_column_width=True)


# In[22]:


from PIL import Image
image_hospital = Image.open('hospital.jpg')


# In[23]:


import streamlit as st
import openpyxl
import base64
csv_exporter=openpyxl.Workbook()
sheet=csv_exporter.active
sheet.cell(row=1,column=1).value='eGFR'
sheet.cell(row=1,column=2).value='Age'
sheet.cell(row=1,column=3).value='NIHSS'
sheet.cell(row=1,column=4).value='Albumin'
sheet.cell(row=1,column=5).value='Albumin-to-globulin_ratio'
sheet.cell(row=1,column=6).value='Serum_creatinine'
sheet.cell(row=1,column=7).value='White_blood_cell_count'
sheet.cell(row=1,column=8).value='Blood_neutrophils_count'
sheet.cell(row=1,column=9).value='Fasting_blood_glucose'
sheet.cell(row=1,column=10).value='Collateral_status'
csv_exporter.save('for predictions.csv')#注意！文件此时保存在内存中且为字节格式文件
data=open('for predictions.csv','rb').read()#以只读模式读取且读取为二进制文件
b64 = base64.b64encode(data).decode('UTF-8')#解码并加密为base64
href = f'<a href="data:file/data;base64,{b64}" download="myresults.csv">Download csv file</a>'#定义下载链接，默认的下载文件名是myresults.xlsx
st.markdown(href, unsafe_allow_html=True)#输出到浏览器
csv_exporter.close()


# In[24]:


file_upload = st.file_uploader("Upload csv file for predictions", type=["csv"])


# In[25]:


if file_upload is not None:
            data = pd.read_csv(file_upload,sep=',')                       
            predictions = predict(model,data) 
            predictions = pd.DataFrame(predictions,columns = ['Predictions'])
            st.write(predictions)


# In[ ]:




