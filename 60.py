#!/usr/bin/env python
# coding: utf-8

# In[1]:


import streamlit as st
import pandas as pd
import numpy as np
import pickle
import shap
shap.initjs()
from sklearn.preprocessing import OneHotEncoder
from PIL import Image
from pycaret import classification
from pycaret.classification import setup, get_logs,create_model, predict_model, tune_model, save_model,load_model


# In[5]:


import streamlit.components.v1 as components


# In[2]:

model = pickle.load(open( "model.p", "rb" ))
model2 = load_model('rf9zzz')

# In[3]:
st.markdown(
    f"""
    <style>
        .appview-container .main .block-container{{
            max-width: {1150}px;
            padding-top: {5}rem;
            padding-right: {10}rem;
            padding-left: {10}rem;
            padding-bottom: {5}rem;
        }}
        .reportview-container .main {{
            color: white;
            background-color: black;
        }}
    </style>
    """
    ,unsafe_allow_html=True,
)



def transform(data):    
        data['Collateral_status'] = pd.factorize(data['Collateral_status'])[0]
        return data


# In[4]:


def predict(model, input_df):
    transform_data = transform(input_df)
    predictions_df = model.predict(transform_data)
    predictions = predictions_df[0]
    return predictions_df

explainer = shap.Explainer(model)

def load_data():
    return shap.datasets.boston()



def st_shap(plot, height=None):
    shap_html = f"<head>{shap.getjs()}</head><body>{plot.html()}</body>"
    components.html(shap_html, height=height)

# In[12]:


def run():


  st.title('Functional outcome prediction  App for patients with  anterior circulation large vessel occlusion after mechanical thrombectomy')

add_selectbox = st.sidebar.selectbox(
    "How would you like to predict?",
    ("Online_number","Batch", "Online_slide"))


# In[13]:


        
image = Image.open('hospital1.webp')
st.image(image, use_column_width=True)


# In[14]:







# In[15]:


if add_selectbox == 'Online_number':
      st.write('This is a web app to predict 90-day outcome of acute ischaemic stroke patients with mechanical thrombectomy based on several features. please fill in the blanks with corresponding data. After that,click on the Predict button at the bottom to see the prediction of the classifier.     Note: all the required laboratory indices are postoperative indices. If the collateral circulation is good, fill in "1"; otherwise, fill in "0".')
    
      Age = st.number_input('Age', min_value = 18,max_value = 95 ,value = 73) 
      NIHSS = st.number_input('NIHSS', min_value = 4,max_value = 38,value = 30)
      CRP = st.number_input('CRP', min_value = 0.20,max_value = 80.00,value = 39.20)
      Albumin = st.number_input('Albumin', min_value = 0.20,max_value = 80.00,value = 39.20)
      AGR = st.number_input('AGR', min_value = 0.50,max_value = 10.00 , value = 1.50)
      Neutrophils = st.number_input('Neutrophils', min_value = 2.00, max_value = 20.00 ,value = 7.40)
      Serum_glucose = st.number_input('Serum_glucose', min_value = 2.50, max_value = 25.00 , value = 11.78)
      eGFR = st.number_input('eGFR', min_value = 10.00,max_value = 250.00,value = 111.5)
      Collateral_status = st.selectbox('Collateral_status', [1, 0])
      output=""


      features_df  = {'Age': Age, 'NIHSS': NIHSS, 
                   'CRP': CRP,'Albumin': Albumin,
                   'AGR': AGR,'Neutrophils': Neutrophils,
                   'Serum_glucose': Serum_glucose, 'eGFR': eGFR,
                  'Collateral_status_0.0': 1-Collateral_status,'Collateral_status_1.0': Collateral_status,'Collateral_status_not_available': 0,}
      print(features_df)


# In[14]:


      input = pd.DataFrame([features_df])
      print(input)


# In[15]:


      shap_values = explainer.shap_values(input)


# In[16]:


      shap.force_plot(explainer.expected_value[1], shap_values[1],input)


# In[17]:

# In[18]:


      if st.button('Predict'): 
            output = model.predict_proba(input)[:,1] 
      st.write(' Based on feature values, predicted possibility of good functional outcome is '+ str(output))
      st_shap(shap.force_plot(explainer.expected_value[1], shap_values[1],input))

# In[16]:


if add_selectbox == 'Batch':
        st.write('This is a web app to predict 90-day outcome of acute ischaemic stroke patients with mechanical thrombectomy based on several features. Please click on the link to download the form and fill in the corresponding data. After that, click on the Browse files button to upload file for prediciton, you can see the prediction of the classifier at the bottom. This page supports batch prediction of the outcome of multiple patients at one time, and can predict the outcome of patients with missing values. Note: all the required laboratory indices are postoperative indices. If the collateral circulation is good, fill in "1"; otherwise, fill in "0". ')
        
        import streamlit as st
        import openpyxl
        import base64
        csv_exporter=openpyxl.Workbook()
        sheet=csv_exporter.active
        sheet.cell(row=1,column=1).value='Age'
        sheet.cell(row=1,column=2).value='NIHSS'
        sheet.cell(row=1,column=3).value='CRP'
        sheet.cell(row=1,column=4).value='Albumin'
        sheet.cell(row=1,column=5).value='AGR'
        sheet.cell(row=1,column=6).value='Neutrophils'
        sheet.cell(row=1,column=9).value='Collateral_status'
        sheet.cell(row=1,column=7).value='Serum_glucose'
        sheet.cell(row=1,column=8).value='eGFR'
        csv_exporter.save('for predictions.csv')#注意！文件此时保存在内存中且为字节格式文件
        data=open('for predictions.csv','rb').read()#以只读模式读取且读取为二进制文件
        b64 = base64.b64encode(data).decode('UTF-8')#解码并加密为base64
        href = f'<a href="data:file/data;base64,{b64}" download="myresults.csv">Download csv file</a>'#定义下载链接，默认的下载文件名是myresults.xlsx
        st.markdown(href, unsafe_allow_html=True)#输出到浏览器
        csv_exporter.close()

        error_bad_lines=False
        file_upload = st.file_uploader("Upload csv file for predictions", type=["csv"])
        
        if file_upload is not None:
            data = pd.read_csv(file_upload,sep=',',error_bad_lines=False)                       
            predictions = model2.predict_proba(data)[:,1] 
            predictions = pd.DataFrame(predictions,columns = ['Predictions'])
            st.write(predictions)


# In[27]:


if add_selectbox == 'Online_slide':
    
      st.write('This is a web app to predict 90-day outcome of acute ischaemic stroke patients with mechanical thrombectomy based on several features that you can see in the sidebar. Please adjust the value of each feature. After that, click on the Predict button at the bottom to see the prediction of the classifier. Note: all the required laboratory indices are postoperative indices. If the collateral circulation is good, fill in "1"; otherwise, fill in "0".')

      Age = st.sidebar.slider(label = 'Age', min_value = 18,
                          max_value = 95 ,
                          value = 73,
                          step = 1) 
      NIHSS = st.sidebar.slider(label = 'NIHSS', min_value = 4,
                          max_value = 38,
                          value = 30,
                          step = 1)
      CRP = st.sidebar.slider(label = 'CRP', min_value = 0.20,
                          max_value = 80.00,
                          value = 39.20,
                          step = 0.10)
      Albumin = st.sidebar.slider(label = 'Albumin', min_value = 0.20,
                          max_value = 80.00,
                          value = 39.20,
                          step = 0.01)
      AGR = st.sidebar.slider(label = 'AGR', min_value = 0.50,
                          max_value = 10.00 ,
                          value = 1.50,
                          step = 0.01)
      Serum_glucose = st.sidebar.slider(label = 'Serum_glucose', min_value = 2.50,
                          max_value = 25.00 ,
                          value = 11.78,
                          step = 0.10) 
      Neutrophils = st.sidebar.slider(label = 'Neutrophils', min_value = 2.00,
                          max_value = 20.00 ,
                          value = 7.40,
                          step = 0.01)
      eGFR = st.sidebar.slider(label = 'eGFR', min_value = 10.00,
                          max_value = 250.00,
                          value = 111.5,
                          step = 0.10)
      Collateral_status = st.sidebar.selectbox('Collateral_status', [1, 0])

      output=""


      features_df  = {'Age': Age, 'NIHSS': NIHSS, 
                   'CRP': CRP,'Albumin': Albumin,
                   'AGR': AGR,'Neutrophils': Neutrophils,
                   'Serum_glucose': Serum_glucose, 'eGFR': eGFR,
                  'Collateral_status_0.0': 1-Collateral_status,'Collateral_status_1.0': Collateral_status,'Collateral_status_not_available': 0,}
      print(features_df)
        
      features = {'Age': Age, 'NIHSS': NIHSS, 
                   'CRP': CRP,'Albumin': Albumin,
                   'AGR': AGR,'Neutrophils': Neutrophils,
                   'Serum_glucose': Serum_glucose, 'eGFR': eGFR,
                  'Collateral_status_0.0': 1-Collateral_status,'Collateral_status_1.0': Collateral_status,'Collateral_status_not_available': 0,}
      print(features)


# In[14]:

      input_df= pd.DataFrame([features])
      print(input_df)
      input = pd.DataFrame([features_df])
      print(input)


# In[15]:


      shap_values = explainer.shap_values(input)



      if st.button('Predict'): 
            output = model.predict_proba(input)[:,1] 
      st.write(' Based on feature values, predicted possibility of good functional outcome is '+ str(output))
      st_shap(shap.force_plot(explainer.expected_value[1], shap_values[1],input), 800)

# In[13]:


if __name__ == '__main__':
    run()

